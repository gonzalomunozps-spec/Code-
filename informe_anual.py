# -*- coding: utf-8 -*-
"""
informe_anual.py
================

Modulo OPCIONAL y DESACOPLADO: genera informes ANUALES de una parcela (toda la
campana, no un estado puntual "de ahora") en tres formatos:

  1. generar_informe_anual   -> PDF de BALANCE (divulgativo): resumen narrativo,
     grafica de evolucion, recorrido fenologico, hitos, estado hidrico,
     uniformidad, intervenciones del cuaderno, radar y valoracion general.
  2. generar_informe_tecnico -> PDF TECNICO (apaisado): tablas completas de
     indices por pasada y por mes, variaciones (deltas), estadisticos de
     campana, fenologia por pasada, heterogeneidad, parametros de radar y
     metodologia (formulas de los indices).
  3. generar_excel           -> .xlsx con los indices por pasada y por MES y
     GRAFICAS embebidas (nativas de Excel), mas hojas de fenologia y radar.

Todo se calcula con el motor real del programa (evaluar_parcela, heterogeneidad,
efecto_producto, interpretar_radar). Lo que ya calculan otros modulos -clima de
ERA5, balance hidrico de la comarca y grados-dia- se ENSENA tal cual: el informe
pide el mismo resumen que ve la ficha y lo pinta, no lo recalcula a su manera.
Igual con lo anotado a mano: variedad, recinto SIGPAC, marca de arbolado disperso
y produccion de bascula salen de la base sin tocarlos.

COMO QUITAR ESTA PARTE:
  Basta con BORRAR este fichero. El panel lo importa de forma tolerante
  (try/except): si no existe, el boton "Informe / Exportar" simplemente no
  aparece y el resto del programa sigue igual. No hay interruptor ni
  configuracion que tocar.

Dependencias (ambas OPCIONALES, tolerantes):
  - reportlab  -> los dos PDF. Si falta, DISPONIBLE = False.
  - openpyxl   -> el Excel. Si falta, EXCEL_DISPONIBLE = False.
  El panel consulta esas banderas y avisa de como instalar lo que falte.
"""

import os
from datetime import datetime

# --- motor real del programa (nucleo; estos modulos siempre estan) ---
from interpretacion_fenologica import evaluar_parcela
from contraste_indices import heterogeneidad
import registro_parcela as REG
import almacen as DB
try:
    import sentinel1 as S1
except Exception:
    S1 = None

# Modulos OPCIONALES del programa. El informe ENSENA lo que ellos calculan, pero
# no lo calcula: si alguno se borra, su seccion desaparece del informe y el resto
# sale igual. Ninguno es obligatorio para generar el documento.
try:
    import grados_dia as _GDD
except Exception:
    _GDD = None
try:
    import clima_era5 as _CLIMA
except Exception:
    _CLIMA = None
try:
    import balance_hidrico as _BH
except Exception:
    _BH = None

# --- reportlab: dependencia propia de este modulo (tolerante) ---
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, HRFlowable,
                                    Table, TableStyle)
    from reportlab.graphics.shapes import Drawing, String, Rect
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.widgets.markers import makeMarker
    _RL = True
except Exception:
    _RL = False

DISPONIBLE = _RL
MOTIVO_NO_DISPONIBLE = ("" if _RL else
                        "Falta la libreria 'reportlab'. Instalala con:  pip install reportlab")

# Secciones OPCIONALES del informe de balance, que el usuario puede elegir incluir
# o no (clave interna, etiqueta para la interfaz). El encabezado, el resumen y la
# valoracion final NO se listan: son el esqueleto y van siempre. `secciones=None`
# en el generador = todas incluidas (comportamiento de siempre).
SECCIONES_BALANCE = [
    ("grafica",     "Gráfica de la campaña"),
    ("fenologia",   "Recorrido fenológico"),
    ("hitos",       "Hitos de la campaña"),
    ("hidrico",     "Estado hídrico (NDMI y balance)"),
    ("clima",       "Clima de la campaña (ERA5)"),
    ("gdd",         "Grados-día (integrales térmicas)"),
    ("uniformidad", "Uniformidad de la parcela"),
    ("cuaderno",    "Intervenciones del cuaderno"),
    ("rendimiento", "Producción registrada"),
    ("progresion",  "Progresión del estado"),
    ("radar",       "Corroboración con radar"),
]

# --- openpyxl: solo para la exportacion a Excel (tolerante) ---
try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XL = True
except Exception:
    _XL = False

EXCEL_DISPONIBLE = _XL
MOTIVO_EXCEL = ("" if _XL else
                "Falta la libreria 'openpyxl'. Instalala con:  pip install openpyxl")

MESES = {1: "enero", 2: "febrero", 3: "marzo", 4: "abril", 5: "mayo", 6: "junio",
         7: "julio", 8: "agosto", 9: "septiembre", 10: "octubre", 11: "noviembre",
         12: "diciembre"}


# =====================================================================
# Helpers puros
# =====================================================================
def _fnat(iso):
    try:
        d = datetime.strptime(iso, "%Y-%m-%d")
        return f"{d.day} de {MESES[d.month]}"
    except (TypeError, ValueError):
        return iso or "-"


def _superficie_ha(coords):
    # geometria compartida en geo.py; aqui se conserva el contrato del informe:
    # None si el poligono no es valido, y hectareas redondeadas en caso contrario.
    if not coords or len(coords) < 3:
        return None
    import geo
    return round(geo.superficie_ha(coords), 2)


def _spec_de(cultivo):
    # mismo modelo de cultivo que usa el panel, centralizado en cultivo.py
    import cultivo as _CU
    return _CU.spec_de(cultivo)


def _eventos_cerca(eventos, fecha_iso, ventana=20):
    """Replica local (sin BD) de eventos en los `ventana` dias previos a fecha_iso."""
    out = []
    if not fecha_iso:
        return out
    for e in eventos or []:
        f = e.get("fecha")
        if not f:
            continue
        try:
            d = (datetime.strptime(fecha_iso, "%Y-%m-%d")
                 - datetime.strptime(f, "%Y-%m-%d")).days
        except (TypeError, ValueError):
            continue
        if 0 <= d <= ventana:
            out.append((d, e))
    return sorted(out, key=lambda x: x[0])


def _num(serie, clave):
    return [r for r in serie if r.get(clave) is not None]


# indices opticos en el orden de presentacion (clave interna -> etiqueta)
INDICES = ["ndvi", "evi", "savi", "gndvi", "lai", "msavi", "ndmi"]
INDICES_ET = {"ndvi": "NDVI", "evi": "EVI", "savi": "SAVI", "gndvi": "GNDVI",
              "lai": "LAI", "msavi": "MSAVI", "ndmi": "NDMI"}


def _agregado_mensual(serie):
    """Media de cada indice por mes de la campana. Devuelve lista ordenada de dicts
    {anio, mes, label, n, <indice>: media, ...}."""
    grupos = {}
    for r in serie:
        try:
            d = datetime.strptime(r["fecha"], "%Y-%m-%d")
        except (TypeError, ValueError):
            continue
        grupos.setdefault((d.year, d.month), []).append(r)
    filas = []
    for (anio, mes) in sorted(grupos):
        regs = grupos[(anio, mes)]
        fila = {"anio": anio, "mes": mes, "label": f"{MESES[mes].capitalize()} {anio}",
                "n": len(regs)}
        for k in INDICES:
            vals = [x[k] for x in regs if x.get(k) is not None]
            fila[k] = round(sum(vals) / len(vals), 3) if vals else None
        filas.append(fila)
    return filas


def _estadisticos(serie):
    """min/max/media/amplitud de cada indice a lo largo de la campana."""
    out = {}
    for k in INDICES:
        vals = [r[k] for r in serie if r.get(k) is not None]
        if vals:
            mn, mx = min(vals), max(vals)
            out[k] = {"min": round(mn, 3), "max": round(mx, 3),
                      "media": round(sum(vals) / len(vals), 3),
                      "amplitud": round(mx - mn, 3), "n": len(vals)}
    return out


# ---------------------------------------------------------------------
# Datos que el informe ENSENA pero no calcula: vienen de la base o de los
# modulos opcionales. Todo lo de aqui es TOLERANTE a fallos: si el modulo no
# esta, la parcela no existe o la tabla esta vacia, se devuelve vacio y la
# seccion correspondiente simplemente no sale. Un informe nunca debe caerse
# porque falte un dato accesorio.
# ---------------------------------------------------------------------
def _clima_mensual(dias):
    """Resumen del clima MES a MES. La tabla diaria de una campana son ~365 filas:
    ilegible en papel. Se agrega igual que el resto del informe (`_agregado_mensual`):
    la lluvia y la ET0 se SUMAN, las temperaturas se promedian y de las extremas se
    guarda la mas extrema, que es como las lee `clima_era5.resumen`."""
    grupos = {}
    for r in dias or []:
        try:
            d = datetime.strptime(r["fecha"], "%Y-%m-%d")
        except (TypeError, ValueError):
            continue
        grupos.setdefault((d.year, d.month), []).append(r)
    filas = []
    for (anio, mes) in sorted(grupos):
        regs = grupos[(anio, mes)]
        def _v(k):
            return [x[k] for x in regs if x.get(k) is not None]
        lluvia, et0 = _v("lluvia"), _v("et0")
        tmed, tmin, tmax = _v("t_media"), _v("t_min"), _v("t_max")
        filas.append({"label": f"{MESES[mes].capitalize()} {anio}", "dias": len(regs),
                      "lluvia": round(sum(lluvia), 1) if lluvia else None,
                      "et0": round(sum(et0), 1) if et0 else None,
                      "balance": (round(sum(lluvia) - sum(et0), 1)
                                  if lluvia and et0 else None),
                      "t_media": round(sum(tmed) / len(tmed), 1) if tmed else None,
                      "t_min": min(tmin) if tmin else None,
                      "t_max": max(tmax) if tmax else None,
                      "heladas": sum(1 for v in tmin if v <= 0.0)})
    return filas


def _clima_de(nombre, campana):
    """Los dias de clima de la campana, o [] si no hay modulo o no hay datos."""
    if _CLIMA is None:
        return []
    try:
        return _CLIMA.clima_de_parcela(nombre, campana) or []
    except Exception:
        return []


def _hidrico_de(dias):
    """Balance hidrico rodante de la comarca al ULTIMO dia con clima.

    Se calcula sobre los dias YA cargados (`_BH.contexto`), no volviendo a la base
    (`contexto_de_parcela`): es el mismo numero que ensena la ficha y ahorra una
    consulta. None si no hay modulo o no hay dias."""
    if _BH is None or not dias:
        return None
    try:
        return _BH.contexto(dias, dias[-1].get("fecha"))
    except Exception:
        return None


def _gdd_de(nombre, tipo, spec, dias):
    """Resumen de grados-dia de la parcela, o None.

    Se ancla en el ULTIMO dia con clima, exactamente igual que la ficha
    (`ficha_clima_gdd._pintar_gdd`): el informe ensena el MISMO numero que se ve en
    pantalla, no uno recalculado a su manera."""
    if _GDD is None or not spec or not spec.get("integrales_termicas") or not dias:
        return None
    try:
        return _GDD.resumen_parcela(tipo, spec.get("especie"), spec,
                                    dias[-1].get("fecha"), nombre)
    except Exception:
        return None


def _rendimientos_de(nombre):
    """Historico de produccion de la parcela (TODAS las campanas), o []."""
    try:
        return DB.rendimientos(nombre) or []
    except Exception:
        return []


def _texto_sigpac(sig):
    """La referencia SIGPAC en una linea, o '' si no esta completa.

    Se escribe en el orden oficial del recinto (prov/mun/agr/zona/pol/par/rec), que
    es como se busca en el visor y como la pide la Administracion."""
    if not sig:
        return ""
    orden = [("Prov", ""), ("Mun", ""), ("Agr", ""), ("Zona", ""),
             ("Pol", ""), ("Par", ""), ("Rec", "")]
    vals = [str(sig.get(k, "") or "").strip() for k, _ in orden]
    if not all(vals[i] for i in (0, 1, 4, 5, 6)):   # sin prov/mun/pol/par/rec no hay recinto
        return ""
    return " / ".join(vals)


def _analisis(nombre, campana, ficha, cultivo, serie, radar, eventos):
    """Ejecuta el motor real sobre la parcela y devuelve el contexto comun que
    consumen los tres formatos (PDF balance, PDF tecnico y Excel)."""
    tipo = (cultivo or {}).get("tipo", "BARBECHO")
    sub = (cultivo or {}).get("subtipo", "")
    especie = (cultivo or {}).get("especie") or tipo.capitalize()
    spec = _spec_de(cultivo)
    coords = (ficha or {}).get("coordenadas")
    superficie = _superficie_ha(coords)
    propietario = (ficha or {}).get("propietario", "-")
    # VARIEDAD: hasta ahora el informe solo decia la especie, y la variedad es lo
    # primero que se pregunta al comparar dos campanas del mismo cultivo.
    variedad = (spec or {}).get("variedad") or ""
    # ARBOLADO DISPERSO: no es un adorno. Con la casilla marcada el diagnostico se
    # calcula sobre el NDVI del CULTIVO (sin las encinas), asi que el informe tiene
    # que decirlo o sus numeros no se pueden reproducir.
    arbolado = bool((ficha or {}).get("arbolado"))
    sigpac = _texto_sigpac((ficha or {}).get("sigpac"))
    rendimientos = _rendimientos_de(nombre)
    # clima -> balance hidrico -> grados-dia: los tres salen de los MISMOS dias,
    # cargados UNA vez.
    clima = _clima_de(nombre, campana)
    clima_resumen = None
    if clima and _CLIMA is not None:
        try:
            clima_resumen = _CLIMA.resumen(clima)
        except Exception:
            clima_resumen = None
    hidrico = _hidrico_de(clima)
    gdd = _gdd_de(nombre, tipo, spec, clima)

    ndvi_serie = _num(serie, "ndvi")
    lai_serie = _num(serie, "lai")
    ndmi_serie = _num(serie, "ndmi")
    inicio, fin = serie[0], serie[-1]
    pico_ndvi = max(ndvi_serie, key=lambda r: r["ndvi"]) if ndvi_serie else None
    pico_lai = max(lai_serie, key=lambda r: r["lai"]) if lai_serie else None
    min_ndmi = min(ndmi_serie, key=lambda r: r["ndmi"]) if ndmi_serie else None

    # --- recorrido fenologico y alertas (motor real, pasada a pasada) ---
    recorrido = []
    avisos = []                      # [(indice, fecha, estado, fase)]
    for i in range(len(serie)):
        parcial = serie[:i + 1]
        evc = _eventos_cerca(eventos, serie[i].get("fecha"))
        d = evaluar_parcela(tipo, sub, parcial, eventos_cerca=evc or None, spec=spec)
        recorrido.append({"fecha": serie[i]["fecha"], "fase": d.get("fase", "-"),
                          "estado": d.get("estado", "-"), "esperado": d.get("esperado", False),
                          "ndvi": serie[i].get("ndvi"), "lai": serie[i].get("lai")})
        if d.get("estado") in ("Revisar", "Vigilar") and not d.get("esperado"):
            avisos.append((i, serie[i]["fecha"], d.get("estado"), d.get("fase")))

    # BALANCE RETROSPECTIVO: un aviso puntual queda "resuelto" si una pasada
    # POSTERIOR lo reencuadra en la fenologia (OK o caida esperada). Solo los que
    # siguen vigentes al cierre pesan en la valoracion general.
    alertas = []                     # [(fecha, estado, fase, resuelto)]
    for (i, fecha, estado, fase) in avisos:
        resuelto = any(recorrido[j]["esperado"] or recorrido[j]["estado"] == "OK"
                       for j in range(i + 1, len(recorrido)))
        alertas.append((fecha, estado, fase, resuelto))
    alertas_vigentes = [a for a in alertas if not a[3]]

    fases_orden = []
    for r in recorrido:
        if r["fase"] and (not fases_orden or fases_orden[-1][0] != r["fase"]):
            fases_orden.append((r["fase"], r["fecha"], r["ndvi"], r["lai"]))

    diag_final = evaluar_parcela(tipo, sub, serie, spec=spec)
    hetero = heterogeneidad(serie)

    # efecto de productos del cuaderno (herbicidas y demas)
    efectos = []
    for e in eventos:
        if e.get("tipo") == "PRODUCTO":
            ef = REG.efecto_producto(serie, e)
            if ef and ef.get("disponible"):
                efectos.append((e, ef))

    # radar
    radar_info = None
    if radar and S1 is not None:
        try:
            radar_info = S1.interpretar_radar(serie, radar, diag_final)
        except Exception:
            radar_info = None

    return {
        "nombre": nombre, "campana": campana, "ficha": ficha, "cultivo": cultivo,
        "serie": serie, "radar": radar, "eventos": eventos,
        "tipo": tipo, "sub": sub, "especie": especie, "spec": spec,
        "superficie": superficie, "propietario": propietario,
        "variedad": variedad, "arbolado": arbolado, "sigpac": sigpac,
        "rendimientos": rendimientos,
        "clima": clima, "clima_resumen": clima_resumen,
        "clima_mensual": _clima_mensual(clima), "hidrico": hidrico, "gdd": gdd,
        "inicio": inicio, "fin": fin,
        "pico_ndvi": pico_ndvi, "pico_lai": pico_lai, "min_ndmi": min_ndmi,
        "recorrido": recorrido, "fases_orden": fases_orden,
        "alertas": alertas, "alertas_vigentes": alertas_vigentes,
        "diag_final": diag_final, "hetero": hetero, "efectos": efectos,
        "radar_info": radar_info,
        "mensual": _agregado_mensual(serie), "estadisticos": _estadisticos(serie),
    }


def texto_progresion_estado(recorrido, alertas_vigentes):
    """Narra COMO evoluciono el estado a lo largo de la campana, en prosa.

    Sustituye a la lista de estado pasada-a-pasada en el PDF: el balance cuenta la
    PROGRESION -de que estado partio, cuanto tiempo estuvo bien, cuando y en que
    fase saltaron los avisos, como cerro-. El detalle dia a dia se deja para el
    Excel, que es donde se consulta un valor concreto. Devuelve HTML de reportlab
    (con &nbsp; y entidades ya escapadas para las tildes de las fases NO, esas van
    aparte)."""
    if not recorrido:
        return ""
    n = len(recorrido)
    cuenta = {"OK": 0, "Vigilar": 0, "Revisar": 0, "Segado": 0}
    for r in recorrido:
        cuenta[r.get("estado")] = cuenta.get(r.get("estado"), 0) + 1
    ini, fin = recorrido[0], recorrido[-1]
    verde = cuenta.get("OK", 0) + cuenta.get("Segado", 0)
    pct_verde = round(100 * verde / n) if n else 0

    frases = [f"A lo largo de {n} pasada(s), la parcela estuvo en un estado sin "
              f"incidencias el {pct_verde}% del seguimiento."]
    frases.append(f"Arranco la campana en «{ini.get('estado','-')}» "
                  f"(fase de {ini.get('fase','-')}) y la cerro en "
                  f"«{fin.get('estado','-')}» (fase de {fin.get('fase','-')}).")

    # los avisos, agrupados por fase, contados (no listados uno a uno)
    avisos = [r for r in recorrido if r.get("estado") in ("Revisar", "Vigilar")
              and not r.get("esperado")]
    if not avisos:
        frases.append("No hubo ningun aviso fuera de lo que la fenologia explica: "
                      "evolucion coherente de principio a fin.")
    else:
        por_fase = {}
        for r in avisos:
            por_fase[r.get("fase", "-")] = por_fase.get(r.get("fase", "-"), 0) + 1
        detalle = ", ".join(f"{c} en {f}" for f, c in por_fase.items())
        frases.append(f"Se registraron {len(avisos)} aviso(s) durante el ano ({detalle}).")
        if alertas_vigentes:
            frases.append(f"De ellos, {len(alertas_vigentes)} seguian vigentes al cierre; "
                          f"el resto se reencuadro en la fenologia en pasadas posteriores.")
        else:
            frases.append("Todos se reencuadraron en la fenologia en pasadas posteriores: "
                          "ninguno quedo vigente al cierre.")
    return " ".join(frases)


def _preparar(serie, radar, eventos, ruta_salida, ext, nombre, campana):
    """Validacion y normalizacion comun de entradas para todos los formatos."""
    serie = sorted([r for r in (serie or []) if r.get("fecha")], key=lambda r: r["fecha"])
    if not serie:
        raise RuntimeError("La parcela no tiene pasadas de satelite: no hay campana que resumir.")
    radar = sorted([r for r in (radar or []) if r.get("fecha")], key=lambda r: r["fecha"])
    eventos = eventos or []
    if not ruta_salida:
        ruta_salida = os.path.abspath(f"Informe_{nombre}_{campana}.{ext}")
    return serie, radar, eventos, ruta_salida


# =====================================================================
# API PUBLICA
# =====================================================================
def secciones_con_datos(nombre, campana, cultivo=None, radar=None, eventos=None):
    """Que secciones del balance tienen datos DE VERDAD para esta parcela.

    Devuelve {clave: True/False} para las claves que se pueden quedar vacias. La
    interfaz lo usa para marcarlas como «(sin datos)» en vez de dejar que el usuario
    pida una seccion que va a salir en blanco. Las que no aparecen aqui se dan por
    disponibles (siempre hay serie: sin ella no hay informe).

    Tolerante: si algo falla, esa clave se da por disponible y como mucho la seccion
    saldra vacia -nunca impide abrir el dialogo-."""
    spec = _spec_de(cultivo)
    try:
        clima = bool(_clima_de(nombre, campana))
    except Exception:
        clima = True
    try:
        rend = bool(_rendimientos_de(nombre))
    except Exception:
        rend = True
    return {"radar": bool(radar), "cuaderno": bool(eventos),
            "clima": clima,
            # el GDD necesita integrales definidas Y clima al que acumular
            "gdd": bool(_GDD is not None and spec and spec.get("integrales_termicas") and clima),
            "rendimiento": rend}



def generar_informe_anual(nombre, campana, ficha, cultivo, serie,
                          radar=None, eventos=None, ruta_salida=None, secciones=None):
    """Informe de BALANCE (PDF, divulgativo). Devuelve la ruta. Lanza RuntimeError.

    `secciones`: iterable con las claves de SECCIONES_BALANCE que se quieren incluir
    (None = todas, como siempre). El encabezado, el resumen y la valoracion final
    van siempre."""
    if not _RL:
        raise RuntimeError(MOTIVO_NO_DISPONIBLE)
    serie, radar, eventos, ruta_salida = _preparar(serie, radar, eventos, ruta_salida,
                                                   "pdf", nombre, campana)
    ctx = _analisis(nombre, campana, ficha, cultivo, serie, radar, eventos)
    _construir_pdf(ruta_salida, ctx, secciones)
    return ruta_salida


def generar_informe_tecnico(nombre, campana, ficha, cultivo, serie,
                            radar=None, eventos=None, ruta_salida=None):
    """Informe TECNICO (PDF): tablas completas de indices por pasada y por mes,
    deltas, estadisticos, parametros de radar y metodologia. Lanza RuntimeError."""
    if not _RL:
        raise RuntimeError(MOTIVO_NO_DISPONIBLE)
    serie, radar, eventos, ruta_salida = _preparar(serie, radar, eventos, ruta_salida,
                                                   "pdf", nombre, campana)
    ctx = _analisis(nombre, campana, ficha, cultivo, serie, radar, eventos)
    _construir_pdf_tecnico(ruta_salida, ctx)
    return ruta_salida


def generar_excel(nombre, campana, ficha, cultivo, serie,
                  radar=None, eventos=None, ruta_salida=None):
    """Hoja de calculo (.xlsx): indices por pasada y por mes, radar y GRAFICAS
    embebidas (nativas de Excel). Requiere openpyxl. Lanza RuntimeError."""
    if not _XL:
        raise RuntimeError(MOTIVO_EXCEL)
    serie, radar, eventos, ruta_salida = _preparar(serie, radar, eventos, ruta_salida,
                                                   "xlsx", nombre, campana)
    ctx = _analisis(nombre, campana, ficha, cultivo, serie, radar, eventos)
    _construir_excel(ruta_salida, ctx)
    return ruta_salida


# =====================================================================
# Construccion del PDF
# =====================================================================
def _construir_pdf(ruta, ctx, secciones=None):
    # que secciones opcionales entran (None = todas, comportamiento de siempre)
    _sec = None if secciones is None else set(secciones)
    def inc(clave):
        return _sec is None or clave in _sec
    HEADER = colors.HexColor("#1e3a2b"); PRIMARY = colors.HexColor("#2f855a")
    PRIMARY_DK = colors.HexColor("#276749"); INK = colors.HexColor("#1a202c")
    INK2 = colors.HexColor("#4a5568"); INK3 = colors.HexColor("#718096")
    BORDER = colors.HexColor("#e2e8f0"); PANEL = colors.HexColor("#f4f8f5")
    COL_NDVI = colors.HexColor("#2f855a"); COL_LAI = colors.HexColor("#dd6b20")
    COL_NDMI = colors.HexColor("#3182ce"); COL_RVI = colors.HexColor("#805ad5")

    serie = ctx["serie"]; radar = ctx["radar"]
    nombre = ctx["nombre"]; campana = ctx["campana"]; especie = ctx["especie"]
    propietario = ctx["propietario"]; superficie = ctx["superficie"]
    cultivo = ctx["cultivo"]; inicio = ctx["inicio"]; fin = ctx["fin"]
    pico_ndvi = ctx["pico_ndvi"]; pico_lai = ctx["pico_lai"]; min_ndmi = ctx["min_ndmi"]
    fases_orden = ctx["fases_orden"]; hetero = ctx["hetero"]; alertas = ctx["alertas"]
    alertas_vigentes = ctx["alertas_vigentes"]
    efectos = ctx["efectos"]; radar_info = ctx["radar_info"]; tipo = ctx["tipo"]
    variedad = ctx["variedad"]; arbolado = ctx["arbolado"]; sigpac = ctx["sigpac"]
    rendimientos = ctx["rendimientos"]; clima_resumen = ctx["clima_resumen"]
    clima_mensual = ctx["clima_mensual"]; hidrico = ctx["hidrico"]; gdd = ctx["gdd"]

    def esc(s): return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    H1 = ParagraphStyle("H1", fontName="Helvetica-Bold", fontSize=13, textColor=PRIMARY_DK,
                        spaceBefore=13, spaceAfter=5, leading=16)
    BODY = ParagraphStyle("BODY", fontName="Helvetica", fontSize=9.5, textColor=INK2,
                          leading=14.5, spaceAfter=5)
    SMALL = ParagraphStyle("SMALL", parent=BODY, fontSize=8.4, textColor=INK3, leading=11.5)
    CELL = ParagraphStyle("CELL", fontName="Helvetica", fontSize=8.6, textColor=INK, leading=11)
    LEAD = ParagraphStyle("LEAD", parent=BODY, fontSize=10.5, leading=15.5)
    WCELL = ParagraphStyle("WCELL", parent=CELL, textColor=colors.white, fontName="Helvetica-Bold")

    def H(t): return Paragraph(t, H1)
    def P(t): return Paragraph(t, BODY)

    def panel(flow, fill=PANEL, bd=BORDER, pad=9):
        t = Table([[flow]], colWidths=[168 * mm])
        t.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), fill),
                               ("BOX", (0, 0), (-1, -1), 0.7, bd),
                               ("TOPPADDING", (0, 0), (-1, -1), pad),
                               ("BOTTOMPADDING", (0, 0), (-1, -1), pad),
                               ("LEFTPADDING", (0, 0), (-1, -1), 11),
                               ("RIGHTPADDING", (0, 0), (-1, -1), 11)]))
        return t

    def grafica():
        n = len(serie); xs = list(range(n)); fechas = [r["fecha"] for r in serie]

        def col(clave, factor=1.0):
            return [(i, serie[i][clave] * factor) for i in range(n) if serie[i].get(clave) is not None]

        # Una linea con MENOS DE DOS puntos hace reventar a reportlab
        # ("Polyline must have 2 or more points"), asi que se descarta. Pasa de
        # verdad: una parcela puede tener NDVI en todas las pasadas y NDMI en
        # ninguna, y entonces el informe entero se caia por una linea vacia.
        lineas = [(et, c, pts) for et, c, pts in
                  (("NDVI (verdor)", COL_NDVI, col("ndvi")),
                   ("LAI /5 (biomasa)", COL_LAI, col("lai", 0.2)),
                   ("NDMI (agua)", COL_NDMI, col("ndmi")))
                  if len(pts) >= 2]
        rvi_pts = []
        if radar:
            for rr in radar:
                if rr.get("rvi") is None:
                    continue
                k = min(range(n), key=lambda i: abs(
                    (datetime.strptime(serie[i]["fecha"], "%Y-%m-%d")
                     - datetime.strptime(rr["fecha"], "%Y-%m-%d")).days))
                rvi_pts.append((k, rr["rvi"]))
        i_rvi = -1
        if len(rvi_pts) >= 2:
            i_rvi = len(lineas)
            lineas.append(("RVI radar (estructura)", COL_RVI, rvi_pts))
        if not lineas:                     # sin ninguna linea no hay grafica que pintar
            return None
        data = [pts for _et, _c, pts in lineas]
        cols = [c for _et, c, _pts in lineas]

        d = Drawing(170 * mm, 76 * mm)
        lp = LinePlot(); lp.x = 14 * mm; lp.y = 15 * mm; lp.width = 150 * mm; lp.height = 52 * mm
        lp.data = data
        for i, c in enumerate(cols):
            lp.lines[i].strokeColor = c; lp.lines[i].strokeWidth = 2 if i == 0 else 1.6
            lp.lines[i].symbol = makeMarker("FilledCircle"); lp.lines[i].symbol.size = 3.2
            lp.lines[i].symbol.fillColor = c
        if i_rvi >= 0:
            lp.lines[i_rvi].strokeDashArray = (3, 2)
        lp.xValueAxis.valueMin = -0.3; lp.xValueAxis.valueMax = n - 0.7; lp.xValueAxis.valueSteps = xs
        lp.xValueAxis.labelTextFormat = lambda v: (fechas[int(round(v))][5:] if 0 <= int(round(v)) < n else "")
        lp.xValueAxis.labels.fontSize = 6.5
        lp.yValueAxis.valueMin = -0.1; lp.yValueAxis.valueMax = 1.0; lp.yValueAxis.valueStep = 0.2
        lp.yValueAxis.labels.fontSize = 6.5
        d.add(lp)
        x = 15 * mm
        for texto, c, _pts in lineas:
            d.add(Rect(x, 70 * mm, 4 * mm, 2.2 * mm, fillColor=c, strokeColor=None))
            d.add(String(x + 5 * mm, 69.6 * mm, esc(texto), fontName="Helvetica", fontSize=7, fillColor=INK2))
            x += 41 * mm
        return d

    story = []
    story.append(Paragraph("Balance de la campa&ntilde;a",
                 ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=20, textColor=HEADER, leading=23)))
    story.append(Paragraph(f"{esc(nombre.replace('_', ' '))} &mdash; {esc(especie)} &middot; Campa&ntilde;a {esc(campana)}",
                 ParagraphStyle("s", fontName="Helvetica", fontSize=12, textColor=PRIMARY, leading=15)))
    story.append(HRFlowable(width="46%", thickness=2, color=PRIMARY, spaceBefore=6, spaceAfter=9, hAlign="LEFT"))

    siembra = (cultivo or {}).get("fecha_siembra")
    cab = [[Paragraph("<b>Propietario</b>", SMALL), Paragraph(esc(propietario), CELL),
            Paragraph("<b>Superficie</b>", SMALL),
            Paragraph(f"{superficie} ha" if superficie else "-", CELL)],
           [Paragraph("<b>Cultivo</b>", SMALL),
            Paragraph(esc(especie) + (f" (siembra {esc(siembra)})" if siembra else ""), CELL),
            Paragraph("<b>Periodo</b>", SMALL),
            Paragraph(f"{_fnat(inicio['fecha'])} &ndash; {_fnat(fin['fecha'])}", CELL)]]
    # Variedad y referencia SIGPAC: identifican el cultivo y el TERRENO. Solo se
    # anaden si constan, para no llenar la cabecera de guiones.
    if variedad or sigpac:
        cab.append([Paragraph("<b>Variedad</b>", SMALL),
                    Paragraph(esc(variedad) if variedad else "-", CELL),
                    Paragraph("<b>Recinto SIGPAC</b>", SMALL),
                    Paragraph(esc(sigpac) if sigpac else "-", CELL)])
    tc = Table(cab, colWidths=[26 * mm, 62 * mm, 24 * mm, 56 * mm])
    tc.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("LINEBELOW", (0, 0), (-1, -1), 0.4, BORDER),
                            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.append(tc)
    # El enmascarado de encinas CAMBIA los numeros del informe, asi que se declara
    # aqui, en el esqueleto: no es una seccion que se pueda desmarcar.
    if arbolado:
        story.append(Paragraph(
            "Parcela marcada como <b>arbolado disperso</b> (dehesa/encinas): el diagn&oacute;stico "
            "y el verdor de este informe se calculan sobre los p&iacute;xeles de CULTIVO, apartando "
            "los de copa. Sin ese filtro la media saldr&iacute;a inflada por el arbolado.", SMALL))
    story.append(Spacer(1, 6))

    # ---- resumen narrativo (data-driven) ----
    frases = [f"A lo largo de la campana {esc(campana)}, la parcela {esc(nombre.replace('_', ' '))} "
              f"({esc(especie)}) fue seguida en {len(serie)} pasadas de satelite entre el "
              f"{_fnat(inicio['fecha'])} y el {_fnat(fin['fecha'])}."]
    if pico_ndvi and inicio.get("ndvi") is not None:
        f = (f"El verdor (NDVI) parti&oacute; de {inicio['ndvi']:.2f}, alcanz&oacute; su m&aacute;ximo de "
             f"{pico_ndvi['ndvi']:.2f} el {_fnat(pico_ndvi['fecha'])}")
        if pico_lai:
            f += f" (con un LAI de {pico_lai['lai']:.1f})"
        if fin.get("ndvi") is not None:
            f += f" y cerr&oacute; la campa&ntilde;a en {fin['ndvi']:.2f}."
        else:
            f += "."
        frases.append(f)
    if min_ndmi:
        frases.append(f"El &iacute;ndice de humedad (NDMI) tuvo su valor m&aacute;s bajo ({min_ndmi['ndmi']:+.2f}) el "
                      f"{_fnat(min_ndmi['fecha'])}.")
    if alertas_vigentes:
        frases.append(f"Al cierre de la campa&ntilde;a quedan {len(alertas_vigentes)} aviso(s) sin reencuadrar "
                      f"por la fenolog&iacute;a (ver m&aacute;s abajo), que conviene revisar.")
    elif alertas:
        frases.append(f"Hubo {len(alertas)} aviso(s) puntual(es) durante el a&ntilde;o, pero las pasadas "
                      f"posteriores los reencuadraron en la fenolog&iacute;a: la campa&ntilde;a cerr&oacute; sin incidencias vigentes.")
    else:
        frases.append("No se detectaron ca&iacute;das fuera de fase, focos localizados ni estr&eacute;s adelantado: "
                      "una evoluci&oacute;n coherente con la fenolog&iacute;a del cultivo de principio a fin.")
    story.append(panel(Paragraph(" ".join(frases), LEAD)))
    story.append(Spacer(1, 4))

    # ---- grafica ----
    graf = grafica() if inc("grafica") else None
    if graf is not None:
        story.append(H("La campa&ntilde;a de un vistazo"))
        story.append(graf)
        story.append(Paragraph("Cada punto es una pasada de sat&eacute;lite. El LAI se muestra dividido por 5 para "
                               "compartir escala. En la aplicaci&oacute;n la gr&aacute;fica es interactiva, con tooltip.", SMALL))

    # ---- recorrido fenologico ----
    if inc("fenologia") and fases_orden:
        story.append(H("Recorrido fenol&oacute;gico"))
        fil = [[Paragraph("Fase", WCELL), Paragraph("Desde", WCELL),
                Paragraph("NDVI", WCELL), Paragraph("LAI", WCELL)]]
        for fase, fecha, ndvi, lai in fases_orden:
            fil.append([Paragraph(esc(fase), CELL), Paragraph(_fnat(fecha), CELL),
                        Paragraph(f"{ndvi:.2f}" if ndvi is not None else "-", CELL),
                        Paragraph(f"{lai:.1f}" if lai is not None else "-", CELL)])
        tf = Table(fil, colWidths=[70 * mm, 44 * mm, 27 * mm, 27 * mm])
        tf.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), PRIMARY_DK),
                                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PANEL]),
                                ("GRID", (0, 0), (-1, -1), 0.4, BORDER), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("ALIGN", (2, 0), (-1, -1), "CENTER"),
                                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(tf)
        story.append(Paragraph("El sistema estima la fase por especie y d&iacute;as desde la siembra; as&iacute; distingue "
                               "una ca&iacute;da propia de la fase (senescencia, corte) de un problema real.", SMALL))

    # ---- hitos ----
    hitos = []
    if pico_ndvi: hitos.append(("M&aacute;ximo verdor (NDVI)", f"{pico_ndvi['ndvi']:.2f}", _fnat(pico_ndvi["fecha"])))
    if pico_lai: hitos.append(("M&aacute;xima biomasa (LAI)", f"{pico_lai['lai']:.1f}", _fnat(pico_lai["fecha"])))
    if min_ndmi: hitos.append(("Momento de menos agua (NDMI)", f"{min_ndmi['ndmi']:+.2f}", _fnat(min_ndmi["fecha"])))
    if fin.get("ndvi") is not None:
        hitos.append(("Cierre de campa&ntilde;a (NDVI)", f"{fin['ndvi']:.2f}", _fnat(fin["fecha"])))
    if inc("hitos") and hitos:
        story.append(H("Hitos de la campa&ntilde;a"))
        hh = [[Paragraph(f"<b>{a}</b>", CELL),
               Paragraph(b, ParagraphStyle("v", parent=CELL, fontName="Helvetica-Bold", textColor=PRIMARY_DK)),
               Paragraph(esc(c), CELL)] for a, b, c in hitos]
        th = Table(hh, colWidths=[78 * mm, 40 * mm, 50 * mm])
        th.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), 0.4, BORDER), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(th)

    # ---- estado hidrico ----
    if inc("hidrico") and (min_ndmi and pico_ndvi or hidrico):
        story.append(H("Estado h&iacute;drico durante el a&ntilde;o"))
        if min_ndmi and pico_ndvi:
            story.append(P(f"El &iacute;ndice de humedad (NDMI) acompa&ntilde;&oacute; al desarrollo del cultivo y toc&oacute; su valor "
                           f"m&aacute;s bajo ({min_ndmi['ndmi']:+.2f}) el {_fnat(min_ndmi['fecha'])}. La firma temprana de un "
                           f"estr&eacute;s h&iacute;drico ser&iacute;a un NDMI que cae ANTES y m&aacute;s r&aacute;pido que el NDVI; el motor la "
                           f"vigila pasada a pasada para poder avisar a tiempo."))
        # BALANCE DE LA COMARCA (lluvia - ET0). Distingue un NDMI bajo por sequia
        # general de uno bajo por un problema DE ESTA parcela, que es justo lo que
        # el lector del informe necesita saber para decidir si actuar.
        if hidrico:
            txt = esc(_BH.texto_contexto(hidrico))
            if hidrico.get("sequia"):
                txt += (" El d&eacute;ficit prolongado explica un NDMI bajo sin que sea, por s&iacute; solo, "
                        "un problema de esta parcela.")
            story.append(P(txt))
            story.append(Paragraph("El balance es de COMARCA: el p&iacute;xel de ERA5-Land mide 11 km de lado, "
                                   "as&iacute; que todas las parcelas de la zona comparten el dato.", SMALL))

    # ---- clima de la campana ----
    if inc("clima") and clima_mensual:
        story.append(H("Clima de la campa&ntilde;a"))
        if clima_resumen:
            story.append(P(esc(_CLIMA.texto_resumen(clima_resumen))))
        fil = [[Paragraph(t, WCELL) for t in ("Mes", "Lluvia mm", "ET0 mm", "Balance mm",
                                              "T media", "T m&iacute;n", "T m&aacute;x", "Heladas")]]
        def _c(v, dec=1, signo=False):
            if v is None:
                return "-"
            return (f"{v:+.{dec}f}" if signo else f"{v:.{dec}f}")
        for f in clima_mensual:
            fil.append([Paragraph(esc(f["label"]), CELL),
                        Paragraph(_c(f["lluvia"], 0), CELL), Paragraph(_c(f["et0"], 0), CELL),
                        Paragraph(_c(f["balance"], 0, True), CELL),
                        Paragraph(_c(f["t_media"]), CELL), Paragraph(_c(f["t_min"]), CELL),
                        Paragraph(_c(f["t_max"]), CELL), Paragraph(str(f["heladas"] or "-"), CELL)])
        tcl = Table(fil, colWidths=[34 * mm] + [17 * mm] * 7)
        tcl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), PRIMARY_DK),
                                 ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PANEL]),
                                 ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                                 ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                 ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                                 ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                                 ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5)]))
        story.append(tcl)
        story.append(Paragraph("Lluvia y ET0 son SUMAS del mes; las temperaturas, medias, salvo "
                               "las extremas, que son la m&aacute;s extrema del mes. Datos de ERA5-Land "
                               "(Copernicus), de comarca y con unos 8 d&iacute;as de retraso.", SMALL))

    # ---- grados-dia ----
    if inc("gdd") and gdd:
        story.append(H("Grados-d&iacute;a (integrales t&eacute;rmicas)"))
        partes = []
        if gdd.get("gdd_acumulado") is not None:
            partes.append(f"Acumulado desde la siembra: <b>{gdd['gdd_acumulado']:.0f} &deg;C&middot;d&iacute;a</b> "
                          f"en {gdd.get('dias', 0)} d&iacute;as"
                          + (f" ({gdd['huecos']} sin dato)" if gdd.get("huecos") else "") + ".")
        if gdd.get("fase_gdd"):
            partes.append(f"Fase por grados-d&iacute;a: <b>{esc(gdd['fase_gdd'])}</b>.")
        if gdd.get("faltan_siguiente") is not None:
            partes.append(f"Faltaban ~{gdd['faltan_siguiente']:.0f} &deg;C&middot;d&iacute;a para la fase siguiente.")
        if not gdd.get("hay_referencia"):
            partes.append("Este cultivo no tiene tabla de referencia de GDD: la fase la marca el calendario.")
        else:
            partes.append("Con integral definida, la fase del diagn&oacute;stico la marca el GDD.")
            if gdd.get("hitos_propios"):
                partes.append("Afinado con TUS valores entre estados, que mandan sobre la bibliograf&iacute;a.")
            if gdd.get("aviso_metodo"):
                partes.append("Aviso: el m&eacute;todo elegido no es &laquo;tiempo t&eacute;rmico&raquo;, as&iacute; que sus unidades no "
                              "coinciden con los hitos de fase y la fase por GDD es orientativa.")
        if partes:
            story.append(P(" ".join(partes)))
        filas_gdd = gdd.get("integrales") or []
        if filas_gdd:
            fil = [[Paragraph(t, WCELL) for t in ("De", "A", "M&eacute;todo", "Referencia", "Origen")]]
            for it in filas_gdd:
                ref = it.get("referencia_gdd")
                fil.append([Paragraph(esc(it.get("desde", "")), CELL),
                            Paragraph(esc(it.get("hasta", "")), CELL),
                            Paragraph(esc(it.get("metodo", "")), CELL),
                            Paragraph(f"{ref:.0f} &deg;C&middot;d&iacute;a" if ref is not None else "-", CELL),
                            Paragraph(esc(it.get("referencia_fuente") or "-"), CELL)])
            tg = Table(fil, colWidths=[33 * mm, 33 * mm, 48 * mm, 27 * mm, 27 * mm])
            tg.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), PRIMARY_DK),
                                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PANEL]),
                                    ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
                                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                    ("TOPPADDING", (0, 0), (-1, -1), 3.5),
                                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3.5)]))
            story.append(tg)
            story.append(Paragraph("Compara la referencia de cada tramo con el acumulado real para ver "
                                   "si el cultivo va adelantado o atrasado.", SMALL))

    # ---- uniformidad ----
    if inc("uniformidad"):
        story.append(H("Uniformidad de la parcela"))
        if hetero and hetero.get("disponible"):
            story.append(P(f"Uniformidad: <b>{esc(hetero.get('uniformidad', '-'))}</b>. {esc(hetero.get('lectura', ''))}"))
        else:
            story.append(P("En las pasadas de esta campa&ntilde;a no se dispuso de estad&iacute;stica espacial interna, por lo "
                           "que el balance se hace sobre los valores medios. Cuando esa estad&iacute;stica est&aacute; disponible, "
                           "el sistema vigila si alg&uacute;n rodal se separa del conjunto (posible foco)."))

    # ---- intervenciones ----
    if inc("cuaderno") and efectos:
        story.append(H("Intervenciones del cuaderno de campo"))
        for e, ef in efectos:
            story.append(P(f"&bull;&nbsp; {esc(_fnat(e.get('fecha')))}: {esc(e.get('producto', '') or e.get('objetivo', ''))} "
                           f"&mdash; {esc(ef.get('verdicto', ''))}."))

    # ---- produccion registrada (bascula, no estimacion) ----
    # Es el unico dato OBJETIVO del sistema y hasta ahora no salia en ningun informe.
    # Se listan TODAS las campanas: el valor de una cosecha esta en compararla.
    if inc("rendimiento") and rendimientos:
        story.append(H("Producci&oacute;n registrada"))
        for r in rendimientos:
            story.append(P("&bull;&nbsp; " + esc(REG.linea_rendimiento(r))))
        story.append(Paragraph("Dato de b&aacute;scula o de parte de cosecha, anotado en el cuaderno: el "
                               "sistema NO lo estima ni lo corrige a humedad comercial. Incluye las "
                               "campa&ntilde;as anteriores que consten.", SMALL))

    # ---- progresion del estado (narrativa, NO lista dia a dia) ----
    # El detalle pasada-a-pasada se deja para el Excel; aqui se cuenta el arco.
    prog = texto_progresion_estado(ctx["recorrido"], alertas_vigentes)
    if inc("progresion") and prog:
        story.append(H("Progresi&oacute;n del estado durante la campa&ntilde;a"))
        story.append(P(esc(prog)))

    # ---- radar ----
    if inc("radar") and radar_info and radar_info.get("disponible"):
        story.append(H("Corroboraci&oacute;n con radar (Sentinel-1)"))
        story.append(P(esc(radar_info.get("texto", ""))))

    # ---- valoracion general ----
    story.append(H("Valoraci&oacute;n general de la campa&ntilde;a"))
    if tipo == "BARBECHO":
        val = (f"Parcela en barbecho durante la campa&ntilde;a {esc(campana)}: no se eval&uacute;a el vigor de un cultivo. "
               "El seguimiento queda como registro del estado del suelo.")
    elif alertas_vigentes:
        val = (f"Campa&ntilde;a con {len(alertas_vigentes)} aviso(s) sin reencuadrar al cierre (ver secci&oacute;n de avisos). "
               f"El cultivo alcanz&oacute; un pico de NDVI de {pico_ndvi['ndvi']:.2f}"
               + (f" y un LAI de {pico_lai['lai']:.1f}" if pico_lai else "")
               + ". Revisa las fechas marcadas y contrasta con lo observado en campo.")
    else:
        val = (f"Campa&ntilde;a <b>favorable y sin incidencias vigentes</b> para {esc(especie)} en "
               f"{esc(nombre.replace('_', ' '))}. La parcela complet&oacute; el ciclo con un pico de vigor "
               f"de NDVI {pico_ndvi['ndvi']:.2f}"
               + (f" y LAI {pico_lai['lai']:.1f}" if pico_lai else "")
               + ". "
               + ("Los avisos puntuales del a&ntilde;o se reencuadraron en la fenolog&iacute;a en pasadas posteriores. "
                  if alertas else "")
               + "Evoluci&oacute;n coherente de principio a fin.")
    story.append(panel(Paragraph(val, LEAD), fill=colors.HexColor("#eef6ef")))

    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=0.6, color=BORDER, spaceBefore=2, spaceAfter=6))
    story.append(Paragraph(f"Informe generado por el Sistema de Gesti&oacute;n y Monitoreo de Parcelas el "
                           f"{datetime.now():%d-%m-%Y %H:%M}. Los &iacute;ndices proceden de Copernicus/Sentinel-2 "
                           f"(y Sentinel-1 si se ha descargado el radar). La interpretaci&oacute;n es orientativa: "
                           f"conviene contrastarla con la observaci&oacute;n en campo.", SMALL))

    def encpie(c, doc):
        c.saveState(); w, h = A4
        c.setStrokeColor(BORDER); c.setLineWidth(0.5); c.line(20 * mm, 14 * mm, w - 20 * mm, 14 * mm)
        c.setFont("Helvetica", 7.5); c.setFillColor(INK3)
        c.drawString(20 * mm, 9.5 * mm, f"Balance de campaña - {nombre.replace('_', ' ')} - {campana}")
        c.drawRightString(w - 20 * mm, 9.5 * mm, f"Pag. {doc.page}")
        c.restoreState()

    doc = SimpleDocTemplate(ruta, pagesize=A4, topMargin=16 * mm, bottomMargin=18 * mm,
                            leftMargin=20 * mm, rightMargin=20 * mm,
                            title=f"Balance de campana - {nombre}")
    doc.build(story, onFirstPage=encpie, onLaterPages=encpie)


# =====================================================================
# PDF TECNICO (tablas completas, deltas, mensual, estadisticos, radar)
# =====================================================================
def _construir_pdf_tecnico(ruta, ctx):
    PRIMARY_DK = colors.HexColor("#276749"); HEADER = colors.HexColor("#1e3a2b")
    INK = colors.HexColor("#1a202c"); INK2 = colors.HexColor("#4a5568")
    INK3 = colors.HexColor("#718096"); BORDER = colors.HexColor("#e2e8f0")
    PANEL = colors.HexColor("#f4f8f5"); ROJO = colors.HexColor("#c53030")
    VERDE = colors.HexColor("#2f855a")
    COL = {"ndvi": colors.HexColor("#2f855a"), "lai": colors.HexColor("#dd6b20"),
           "ndmi": colors.HexColor("#3182ce")}

    serie = ctx["serie"]; radar = ctx["radar"]; nombre = ctx["nombre"]; campana = ctx["campana"]
    especie = ctx["especie"]; propietario = ctx["propietario"]; superficie = ctx["superficie"]
    mensual = ctx["mensual"]; estad = ctx["estadisticos"]
    recorrido = ctx["recorrido"]; hetero = ctx["hetero"]; diag_final = ctx["diag_final"]
    variedad = ctx["variedad"]; arbolado = ctx["arbolado"]; sigpac = ctx["sigpac"]
    rendimientos = ctx["rendimientos"]; clima_resumen = ctx["clima_resumen"]
    clima_mensual = ctx["clima_mensual"]; hidrico = ctx["hidrico"]; gdd = ctx["gdd"]

    def esc(s): return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    H1 = ParagraphStyle("TH1", fontName="Helvetica-Bold", fontSize=12, textColor=PRIMARY_DK,
                        spaceBefore=11, spaceAfter=4, leading=15)
    BODY = ParagraphStyle("TB", fontName="Helvetica", fontSize=8.6, textColor=INK2, leading=12, spaceAfter=3)
    SMALL = ParagraphStyle("TS", parent=BODY, fontSize=7.6, textColor=INK3, leading=10)
    C = ParagraphStyle("TC", fontName="Helvetica", fontSize=7.6, textColor=INK, leading=9.5, alignment=1)
    CL = ParagraphStyle("TCL", parent=C, alignment=0)
    W = ParagraphStyle("TW", parent=C, textColor=colors.white, fontName="Helvetica-Bold")

    _n_sec = [0]

    def SEC(t):
        """Titulo de seccion, numerado SOLO. Antes el numero iba escrito en cada
        titulo, asi que insertar una seccion obligaba a renumerar las siguientes a
        mano (y una se quedaba atras)."""
        _n_sec[0] += 1
        return Paragraph(f"{_n_sec[0]}. {t}", H1)

    def celda(v, fmt="{:.3f}", pct=False):
        if v is None:
            return Paragraph("&ndash;", C)
        try:
            return Paragraph(fmt.format(v), C)
        except Exception:
            return Paragraph(esc(v), C)

    def _num(fila, k):  # media mensual coloreada segun signo si es delta
        return fila.get(k)

    def tabla(cabeceras, filas, anchos, aliniz=()):
        head = [Paragraph(f"<b>{esc(h)}</b>", W) for h in cabeceras]
        data = [head] + filas
        t = Table(data, colWidths=anchos, repeatRows=1)
        est = [("BACKGROUND", (0, 0), (-1, 0), PRIMARY_DK),
               ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PANEL]),
               ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
               ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
               ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]
        t.setStyle(TableStyle(est))
        return t

    def grafica_ancha():
        n = len(serie); xs = list(range(n)); fechas = [r["fecha"] for r in serie]

        def col(clave, factor=1.0):
            return [(i, serie[i][clave] * factor) for i in range(n) if serie[i].get(clave) is not None]
        # misma cautela que en el PDF de balance: una linea de 0 o 1 punto hace
        # reventar a reportlab, y un indice puede faltar en toda la campana
        lineas = [(et, c, pts) for et, c, pts in
                  (("NDVI", COL["ndvi"], col("ndvi")),
                   ("LAI /5", COL["lai"], col("lai", 0.2)),
                   ("NDMI", COL["ndmi"], col("ndmi")))
                  if len(pts) >= 2]
        if not lineas:
            return None
        data = [pts for _et, _c, pts in lineas]
        cols = [c for _et, c, _pts in lineas]
        d = Drawing(250 * mm, 62 * mm)
        lp = LinePlot(); lp.x = 12 * mm; lp.y = 12 * mm; lp.width = 232 * mm; lp.height = 40 * mm
        lp.data = data
        for i, c in enumerate(cols):
            lp.lines[i].strokeColor = c; lp.lines[i].strokeWidth = 1.6
            lp.lines[i].symbol = makeMarker("FilledCircle"); lp.lines[i].symbol.size = 3
            lp.lines[i].symbol.fillColor = c
        lp.xValueAxis.valueMin = -0.3; lp.xValueAxis.valueMax = n - 0.7; lp.xValueAxis.valueSteps = xs
        lp.xValueAxis.labelTextFormat = lambda v: (fechas[int(round(v))][5:] if 0 <= int(round(v)) < n else "")
        lp.xValueAxis.labels.fontSize = 6
        lp.yValueAxis.valueMin = -0.1; lp.yValueAxis.valueMax = 1.0; lp.yValueAxis.valueStep = 0.2
        lp.yValueAxis.labels.fontSize = 6
        d.add(lp)
        x = 14 * mm
        for texto, c, _pts in lineas:
            d.add(Rect(x, 56 * mm, 4 * mm, 2 * mm, fillColor=c, strokeColor=None))
            d.add(String(x + 5 * mm, 55.6 * mm, texto, fontName="Helvetica", fontSize=6.5, fillColor=INK2))
            x += 26 * mm
        return d

    story = []
    story.append(Paragraph("Informe t&eacute;cnico de campa&ntilde;a",
                 ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=17, textColor=HEADER, leading=20)))
    story.append(Paragraph(f"{esc(nombre.replace('_', ' '))} &middot; {esc(especie)}"
                           + (f" &middot; var. {esc(variedad)}" if variedad else "")
                           + f" &middot; Campa&ntilde;a {esc(campana)} &middot; {esc(propietario)}"
                           + (f" &middot; {superficie} ha" if superficie else ""),
                 ParagraphStyle("s", fontName="Helvetica", fontSize=9.5, textColor=INK2, leading=13)))
    pie_id = []
    if sigpac:
        pie_id.append(f"Recinto SIGPAC {esc(sigpac)}")
    if arbolado:
        pie_id.append("arbolado disperso: los &iacute;ndices del diagn&oacute;stico se calculan sobre los "
                      "p&iacute;xeles de cultivo, apartando los de copa")
    if pie_id:
        story.append(Paragraph(" &middot; ".join(pie_id), SMALL))
    story.append(HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#2f855a"),
                            spaceBefore=4, spaceAfter=6))

    # 1. Indices por pasada
    story.append(SEC("&Iacute;ndices de vegetaci&oacute;n por pasada"))
    cab = ["Fecha"] + [INDICES_ET[k] for k in INDICES] + ["Cob.%", "NDVI_std"]
    filas = []
    for r in serie:
        fila = [Paragraph(r["fecha"], C)]
        for k in INDICES:
            fila.append(celda(r.get(k)))
        cob = r.get("cobertura_valida")
        fila.append(Paragraph(f"{int(cob * 100)}" if cob is not None else "&ndash;", C))
        fila.append(celda(r.get("ndvi_std")))
        filas.append(fila)
    anchos = [22 * mm] + [23 * mm] * 7 + [15 * mm, 20 * mm]
    story.append(tabla(cab, filas, anchos))

    # 2. Variacion (delta) entre pasadas consecutivas
    story.append(SEC("Variaci&oacute;n entre pasadas consecutivas (&Delta;)"))
    filas = []
    for i in range(1, len(serie)):
        a, b = serie[i - 1], serie[i]
        fila = [Paragraph(f"{a['fecha']} &rarr; {b['fecha']}", CL)]
        for k in INDICES:
            if a.get(k) is not None and b.get(k) is not None:
                dv = round(b[k] - a[k], 3)
                c = ROJO if dv < 0 else VERDE
                fila.append(Paragraph(f"<font color='{c.hexval()}'>{dv:+.3f}</font>", C))
            else:
                fila.append(Paragraph("&ndash;", C))
        filas.append(fila)
    if filas:
        story.append(tabla(["Intervalo"] + [INDICES_ET[k] for k in INDICES],
                           filas, [40 * mm] + [24.5 * mm] * 7))
    else:
        story.append(Paragraph("Una sola pasada: no hay variaci&oacute;n que calcular.", SMALL))

    # 3. Agregado mensual (medias por mes)
    story.append(SEC("Medias mensuales por &iacute;ndice"))
    filas = []
    for f in mensual:
        fila = [Paragraph(esc(f["label"]), CL), Paragraph(str(f["n"]), C)]
        for k in INDICES:
            fila.append(celda(f.get(k)))
        filas.append(fila)
    story.append(tabla(["Mes", "n"] + [INDICES_ET[k] for k in INDICES],
                       filas, [30 * mm, 12 * mm] + [23.4 * mm] * 7))
    story.append(Paragraph("n = n&uacute;mero de pasadas v&aacute;lidas promediadas en el mes.", SMALL))

    # 3b. Variacion mes a mes (delta de las medias mensuales)
    filas = []
    for i in range(1, len(mensual)):
        a, b = mensual[i - 1], mensual[i]
        fila = [Paragraph(f"{esc(a['label'])} &rarr; {esc(b['label'])}", CL)]
        for k in INDICES:
            if a.get(k) is not None and b.get(k) is not None:
                dv = round(b[k] - a[k], 3)
                cc = ROJO if dv < 0 else VERDE
                fila.append(Paragraph(f"<font color='{cc.hexval()}'>{dv:+.3f}</font>", C))
            else:
                fila.append(Paragraph("&ndash;", C))
        filas.append(fila)
    if filas:
        story.append(Spacer(1, 3))
        story.append(Paragraph("<b>Variaci&oacute;n mes a mes (&Delta; de las medias mensuales)</b>", BODY))
        story.append(tabla(["Intervalo"] + [INDICES_ET[k] for k in INDICES],
                           filas, [46 * mm] + [23.7 * mm] * 7))

    # grafica
    story.append(Spacer(1, 4))
    _graf = grafica_ancha()
    if _graf is not None:
        story.append(_graf)

    # 4. Estadisticos de campana
    story.append(SEC("Estad&iacute;sticos de la campa&ntilde;a"))
    filas = []
    for k in INDICES:
        e = estad.get(k)
        if not e:
            continue
        filas.append([Paragraph(f"<b>{INDICES_ET[k]}</b>", CL), celda(e["min"]), celda(e["max"]),
                      celda(e["media"]), celda(e["amplitud"]), Paragraph(str(e["n"]), C)])
    story.append(tabla(["Índice", "Mín", "Máx", "Media", "Amplitud", "n"],
                       filas, [28 * mm, 24 * mm, 24 * mm, 24 * mm, 26 * mm, 14 * mm]))

    # 5. Fenologia y estado por pasada
    story.append(SEC("Fase fenol&oacute;gica y estado por pasada"))
    filas = []
    for r in recorrido:
        est = r["estado"]
        cc = {"OK": VERDE, "Vigilar": colors.HexColor("#d69e2e"),
              "Revisar": colors.HexColor("#dd6b20"), "Segado": VERDE}.get(est, INK)
        filas.append([Paragraph(r["fecha"], C), Paragraph(esc(r["fase"]), CL),
                      Paragraph(f"<font color='{cc.hexval()}'><b>{esc(est)}</b></font>", C),
                      Paragraph("s&iacute;" if r["esperado"] else "no", C)])
    story.append(tabla(["Fecha", "Fase estimada", "Estado", "Esperado"],
                       filas, [24 * mm, 90 * mm, 30 * mm, 24 * mm]))
    lo, hi = diag_final.get("rango_fase", (None, None))
    story.append(Paragraph(f"Diagn&oacute;stico final ({esc(diag_final.get('fase', '-'))}): "
                           f"<b>{esc(diag_final.get('estado', '-'))}</b>. "
                           + (f"Rango NDVI esperado en la fase: {lo:.2f}&ndash;{hi:.2f}. " if lo is not None else "")
                           + esc(diag_final.get("motivo", "")), SMALL))

    # 6. Heterogeneidad
    story.append(SEC("Distribuci&oacute;n intraparcela (heterogeneidad)"))
    if hetero and hetero.get("disponible"):
        pares = [("Media NDVI", hetero.get("media")), ("Std", hetero.get("std")),
                 ("CV", hetero.get("cv")), ("p10", hetero.get("p10")), ("p50", hetero.get("p50")),
                 ("p90", hetero.get("p90")), ("Amplitud p90-p10", hetero.get("amplitud")),
                 ("&Delta;media", hetero.get("d_media")), ("&Delta;std", hetero.get("d_std"))]
        fila = []
        for et, v in pares:
            fila.append(Paragraph(f"<b>{et}</b><br/>" + (f"{v:.3f}" if isinstance(v, (int, float)) else "&ndash;"), C))
        t = Table([fila], colWidths=[21 * mm] * len(pares))
        t.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.4, BORDER), ("BACKGROUND", (0, 0), (-1, -1), PANEL),
                               ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        story.append(t)
        story.append(Paragraph(f"Uniformidad: <b>{esc(hetero.get('uniformidad', '-'))}</b>. "
                               f"Patr&oacute;n: {esc(hetero.get('patron', '-'))}. {esc(hetero.get('lectura', ''))}", SMALL))
    else:
        story.append(Paragraph("Sin estad&iacute;stica espacial (std/percentiles) en las pasadas de esta campa&ntilde;a.", SMALL))

    # 7. Radar Sentinel-1
    if radar:
        story.append(SEC("Par&aacute;metros de radar (Sentinel-1)"))
        cab = ["Fecha", "VV dB", "VH dB", "RVI", "RVI min", "RVI max", "CR dB", "Fiab.", "Órbita", "n px"]
        filas = []
        for r in radar:
            filas.append([Paragraph(r["fecha"], C), celda(r.get("vv"), "{:.1f}"), celda(r.get("vh"), "{:.1f}"),
                          celda(r.get("rvi")), celda(r.get("rvi_lo")), celda(r.get("rvi_hi")),
                          celda(r.get("cr"), "{:.1f}"), Paragraph(esc(r.get("fiabilidad", "-")), C),
                          Paragraph(esc(r.get("orbita", "-")), C),
                          Paragraph(str(r.get("n_pixeles", "&ndash;")), C)])
        story.append(tabla(cab, filas, [22 * mm, 18 * mm, 18 * mm, 16 * mm, 18 * mm, 18 * mm,
                                        16 * mm, 18 * mm, 18 * mm, 14 * mm]))
        story.append(Paragraph("RVI = 4&middot;VH/(VV+VH) en potencia lineal; RVI min/max = rango por incertidumbre "
                               "(speckle). CR = VH&minus;VV (dB). Fiabilidad heur&iacute;stica por n&ordm; de p&iacute;xeles y dispersi&oacute;n.", SMALL))

    # Clima de la campana (mes a mes; el diario son ~365 filas)
    if clima_mensual:
        story.append(SEC("Clima de la campa&ntilde;a (ERA5-Land, mensual)"))
        if clima_resumen:
            story.append(Paragraph(esc(_CLIMA.texto_resumen(clima_resumen)), BODY))
        def _cc(v, dec=1, signo=False):
            if v is None:
                return Paragraph("&ndash;", C)
            return Paragraph((f"{v:+.{dec}f}" if signo else f"{v:.{dec}f}"), C)
        filas = [[Paragraph(esc(f["label"]), C), Paragraph(str(f["dias"]), C),
                  _cc(f["lluvia"], 0), _cc(f["et0"], 0), _cc(f["balance"], 0, True),
                  _cc(f["t_media"]), _cc(f["t_min"]), _cc(f["t_max"]),
                  Paragraph(str(f["heladas"] or "&ndash;"), C)] for f in clima_mensual]
        story.append(tabla(["Mes", "días", "Lluvia mm", "ET0 mm", "Balance mm",
                            "T med °C", "T mín °C", "T máx °C", "Heladas"],
                           filas, [30 * mm] + [24 * mm] * 8))
        nota = ("Lluvia y ET0 son sumas del mes; las extremas, la más extrema del mes. "
                "El píxel de ERA5-Land mide 11 km de lado: el dato es de COMARCA, no de "
                "parcela, y llega con unos 8 días de retraso.")
        if hidrico:
            nota = _BH.texto_contexto(hidrico) + "  " + nota
        story.append(Paragraph(esc(nota), SMALL))

    # Grados-dia
    if gdd:
        story.append(SEC("Grados-d&iacute;a (integrales t&eacute;rmicas)"))
        res = []
        if gdd.get("gdd_acumulado") is not None:
            res.append(f"Acumulado desde la siembra: {gdd['gdd_acumulado']:.0f} °C·día "
                       f"({gdd.get('dias', 0)} días"
                       + (f", {gdd['huecos']} sin dato" if gdd.get("huecos") else "") + ")")
        if gdd.get("fase_gdd"):
            res.append(f"fase por GDD: {gdd['fase_gdd']}")
        if gdd.get("faltan_siguiente") is not None:
            res.append(f"faltan ~{gdd['faltan_siguiente']:.0f} °C·día para la siguiente")
        res.append(f"método de la fase: {gdd.get('metodo_fase', '-')}")
        if gdd.get("hitos_propios"):
            res.append("hitos propios de la parcela (mandan sobre la bibliografía)")
        if gdd.get("aviso_metodo"):
            res.append("AVISO: el método no es «tiempo térmico», las unidades no casan con los "
                       "hitos de fase y la fase por GDD es orientativa")
        story.append(Paragraph(esc("  ·  ".join(res)), BODY))
        filas = []
        for it in gdd.get("integrales") or []:
            ref = it.get("referencia_gdd")
            filas.append([Paragraph(esc(it.get("desde", "")), CL),
                          Paragraph(esc(it.get("hasta", "")), CL),
                          Paragraph(esc(it.get("metodo", "")), CL),
                          Paragraph(esc(it.get("metodo_clave", "")), C),
                          celda(it.get("cero_vegetativo"), "{:.1f}"),
                          celda(it.get("tope"), "{:.1f}"),
                          Paragraph(f"{ref:.0f}" if ref is not None else "&ndash;", C),
                          Paragraph(esc(it.get("referencia_fuente") or "&ndash;"), C)])
        if filas:
            story.append(tabla(["De", "A", "Método", "Clave", "Cero veg. °C", "Tope °C",
                                "Referencia °C·día", "Origen"],
                               filas, [30 * mm, 30 * mm, 46 * mm, 30 * mm, 24 * mm, 20 * mm,
                                       28 * mm, 24 * mm], aliniz=(0, 1, 2)))

    # Produccion registrada (bascula): el unico dato objetivo, y de TODAS las campanas
    if rendimientos:
        story.append(SEC("Producci&oacute;n registrada (b&aacute;scula)"))
        filas = []
        for r in rendimientos:
            filas.append([Paragraph(esc(r.get("campana", "")), C),
                          Paragraph("Siega" if r.get("tipo") == "SIEGA" else "Cosecha", C),
                          celda(r.get("rendimiento_kg_ha"), "{:,.0f}"),
                          celda(r.get("humedad_grano_pct"), "{:.1f}"),
                          celda(r.get("superficie_cosechada_ha"), "{:.2f}"),
                          Paragraph(esc(r.get("fuente_dato") or "&ndash;"), C),
                          Paragraph(esc(r.get("fecha") or "&ndash;"), C)])
        story.append(tabla(["Campaña", "Tipo", "kg/ha", "Humedad %", "Superficie ha",
                            "Origen", "Fecha"],
                           filas, [30 * mm, 24 * mm, 26 * mm, 24 * mm, 28 * mm, 34 * mm, 26 * mm]))
        story.append(Paragraph("Dato anotado en el cuaderno, no estimado por el sistema: no se "
                               "corrige a humedad comercial ni se normaliza por superficie.", SMALL))

    # Metodologia
    story.append(SEC("Metodolog&iacute;a e &iacute;ndices"))
    metod = [
        "NDVI = (NIR&minus;RED)/(NIR+RED) &mdash; verdor / actividad fotosint&eacute;tica.",
        "GNDVI = (NIR&minus;GREEN)/(NIR+GREEN); NDMI = (NIR&minus;SWIR1)/(NIR+SWIR1) &mdash; humedad del dosel.",
        "SAVI = 1,5&middot;(NIR&minus;RED)/(NIR+RED+0,5); MSAVI corrige el efecto suelo; LAI, &iacute;ndice de &aacute;rea foliar.",
        "Fuente: Copernicus/Sentinel-2 (SR, armonizado) con enmascarado de nubes por SCL; Sentinel-1 (GRD) para el radar.",
        "La fase fenol&oacute;gica se estima por especie y d&iacute;as desde la siembra (le&ntilde;osos por mes y marco). "
        "Una ca&iacute;da propia de la fase (senescencia, siega) no se marca como anomal&iacute;a.",
    ]
    for m in metod:
        story.append(Paragraph("&bull;&nbsp; " + m, SMALL))

    def encpie(c, doc):
        c.saveState(); w, h = landscape(A4)
        c.setStrokeColor(BORDER); c.setLineWidth(0.5); c.line(12 * mm, 10 * mm, w - 12 * mm, 10 * mm)
        c.setFont("Helvetica", 7); c.setFillColor(INK3)
        c.drawString(12 * mm, 6 * mm, f"Informe tecnico - {nombre.replace('_', ' ')} - {campana} - generado {datetime.now():%d-%m-%Y %H:%M}")
        c.drawRightString(w - 12 * mm, 6 * mm, f"Pag. {doc.page}")
        c.restoreState()

    doc = SimpleDocTemplate(ruta, pagesize=landscape(A4), topMargin=13 * mm, bottomMargin=13 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            title=f"Informe tecnico - {nombre}")
    doc.build(story, onFirstPage=encpie, onLaterPages=encpie)


# =====================================================================
# EXCEL (.xlsx) con indices por pasada, por mes y graficas embebidas
# =====================================================================
def _construir_excel(ruta, ctx):
    serie = ctx["serie"]; radar = ctx["radar"]; mensual = ctx["mensual"]
    estad = ctx["estadisticos"]; recorrido = ctx["recorrido"]
    nombre = ctx["nombre"]; campana = ctx["campana"]; especie = ctx["especie"]
    propietario = ctx["propietario"]; superficie = ctx["superficie"]; cultivo = ctx["cultivo"]
    variedad = ctx["variedad"]; arbolado = ctx["arbolado"]; sigpac = ctx["sigpac"]
    rendimientos = ctx["rendimientos"]; clima = ctx["clima"]
    clima_resumen = ctx["clima_resumen"]; gdd = ctx["gdd"]

    CAB = "276749"
    f_cab = Font(bold=True, color="FFFFFF"); relleno_cab = PatternFill("solid", fgColor=CAB)
    f_tit = Font(bold=True, size=14, color="1E3A2B")
    centro = Alignment(horizontal="center")
    borde = Border(*[Side(style="thin", color="E2E8F0")] * 4)

    wb = Workbook()

    def encabeza(ws, cols, fila=1):
        for j, c in enumerate(cols, 1):
            cel = ws.cell(row=fila, column=j, value=c)
            cel.font = f_cab; cel.fill = relleno_cab; cel.alignment = centro; cel.border = borde

    def anchos(ws, ancho, ncols):
        for j in range(1, ncols + 1):
            ws.column_dimensions[get_column_letter(j)].width = ancho

    # ---- Hoja 1: Resumen ----
    ws = wb.active; ws.title = "Resumen"
    ws["A1"] = f"Balance de campaña — {nombre.replace('_',' ')}"; ws["A1"].font = f_tit
    info = [("Parcela", nombre.replace("_", " ")), ("Cultivo", especie),
            ("Variedad", variedad or "-"),
            ("Propietario", propietario), ("Superficie (ha)", superficie or "-"),
            ("Recinto SIGPAC", sigpac or "-"),
            ("Arbolado disperso", "sí (índices sobre píxeles de cultivo)" if arbolado else "no"),
            ("Campaña", campana), ("Siembra", (cultivo or {}).get("fecha_siembra", "-")),
            ("Pasadas", len(serie)), ("Periodo", f"{serie[0]['fecha']} a {serie[-1]['fecha']}")]
    if clima_resumen:
        info.append(("Clima de la campaña", _CLIMA.texto_resumen(clima_resumen)))
    if gdd and gdd.get("gdd_acumulado") is not None:
        info.append(("GDD acumulado", f"{gdd['gdd_acumulado']:.0f} °C·día"
                                      f" ({gdd.get('dias', 0)} días)"))
        if gdd.get("fase_gdd"):
            info.append(("Fase por GDD", gdd["fase_gdd"]))
    for i, (k, v) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 20; ws.column_dimensions["B"].width = 46
    # la tabla de estadisticos arranca DEBAJO de la ficha, que ahora es mas larga
    fila_est = 3 + len(info) + 1
    ws.cell(row=fila_est, column=1, value="Estadísticos de campaña").font = Font(bold=True, size=11)
    encabeza(ws, ["Índice", "Mín", "Máx", "Media", "Amplitud", "n"], fila=fila_est + 1)
    r = fila_est + 2
    for k in INDICES:
        e = estad.get(k)
        if not e:
            continue
        for j, v in enumerate([INDICES_ET[k], e["min"], e["max"], e["media"], e["amplitud"], e["n"]], 1):
            cel = ws.cell(row=r, column=j, value=v); cel.border = borde
            if j > 1:
                cel.alignment = centro
        r += 1

    # ---- Hoja 2: Datos por pasada ----
    # El estado y la fase pasada-a-pasada -la "lista de estado diario"- viven AQUI,
    # no en el PDF: el PDF cuenta la progresion y el Excel guarda el detalle para
    # consultar un dia concreto. El recorrido ya lo trae el ctx (`recorrido`).
    estado_por_fecha = {r["fecha"]: (r.get("estado", "-"), r.get("fase", "-"))
                        for r in recorrido}
    ws = wb.create_sheet("Índices por pasada")
    cols = ["Fecha", "Estado", "Fase"] + [INDICES_ET[k] for k in INDICES] + ["Cobertura %", "NDVI_std"]
    encabeza(ws, cols)
    for i, reg in enumerate(serie, 2):
        ws.cell(row=i, column=1, value=reg["fecha"])
        est, fase = estado_por_fecha.get(reg["fecha"], ("-", "-"))
        ws.cell(row=i, column=2, value=est)
        ws.cell(row=i, column=3, value=fase)
        for j, k in enumerate(INDICES, 4):
            ws.cell(row=i, column=j, value=reg.get(k))
        cob = reg.get("cobertura_valida")
        ws.cell(row=i, column=len(INDICES) + 4, value=round(cob * 100, 1) if cob is not None else None)
        ws.cell(row=i, column=len(INDICES) + 5, value=reg.get("ndvi_std"))
    anchos(ws, 12, len(cols))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10; ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A2"
    # grafica: NDVI, LAI, NDMI por pasada
    ch = LineChart(); ch.title = "Evolución de índices por pasada"; ch.height = 9; ch.width = 22
    ch.y_axis.title = "valor índice"; ch.x_axis.title = "fecha"
    fechas_ref = Reference(ws, min_col=1, min_row=2, max_row=len(serie) + 1)
    # +2 columnas por Estado y Fase, que ahora van antes de los indices
    for k, colidx in (("ndvi", 4), ("lai", 8), ("ndmi", 10)):
        data = Reference(ws, min_col=colidx, min_row=1, max_row=len(serie) + 1)
        ch.add_data(data, titles_from_data=True)
    ch.set_categories(fechas_ref)
    ws.add_chart(ch, "L2")

    # ---- Hoja 3: Medias mensuales ----
    ws = wb.create_sheet("Medias mensuales")
    cols = ["Mes", "n"] + [INDICES_ET[k] for k in INDICES]
    encabeza(ws, cols)
    for i, f in enumerate(mensual, 2):
        ws.cell(row=i, column=1, value=f["label"])
        ws.cell(row=i, column=2, value=f["n"])
        for j, k in enumerate(INDICES, 3):
            ws.cell(row=i, column=j, value=f.get(k))
    anchos(ws, 12, len(cols)); ws.column_dimensions["A"].width = 16
    ws.freeze_panes = "A2"
    if len(mensual) >= 1:
        ch = LineChart(); ch.title = "Medias mensuales por índice"; ch.height = 9; ch.width = 22
        ch.y_axis.title = "media mensual"; ch.x_axis.title = "mes"
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(mensual) + 1)
        for k, colidx in (("ndvi", 3), ("lai", 7), ("ndmi", 9)):
            data = Reference(ws, min_col=colidx, min_row=1, max_row=len(mensual) + 1)
            ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ws.add_chart(ch, "M2")

    # ---- Hoja 3b: Variación mensual (deltas de las medias mensuales) ----
    if len(mensual) >= 2:
        ws = wb.create_sheet("Variación mensual")
        cols = ["Intervalo"] + [INDICES_ET[k] for k in INDICES]
        encabeza(ws, cols)
        rojo = Font(color="C53030"); verde = Font(color="2F855A")
        for i in range(1, len(mensual)):
            a, b = mensual[i - 1], mensual[i]
            fila = i + 1                                   # cabecera en fila 1
            ws.cell(row=fila, column=1, value=f"{a['label']} → {b['label']}")
            for j, k in enumerate(INDICES, 2):
                if a.get(k) is not None and b.get(k) is not None:
                    dv = round(b[k] - a[k], 3)
                    cel = ws.cell(row=fila, column=j, value=dv)
                    cel.font = rojo if dv < 0 else verde
                    cel.alignment = centro
        anchos(ws, 12, len(cols)); ws.column_dimensions["A"].width = 26
        ws.freeze_panes = "A2"
        # grafica de barras: variacion mensual de NDVI, LAI, NDMI
        ch = BarChart(); ch.type = "col"; ch.title = "Variación mes a mes (Δ medias)"
        ch.height = 9; ch.width = 22; ch.y_axis.title = "Δ índice"; ch.x_axis.title = "intervalo"
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(mensual))
        for k, colidx in (("ndvi", 2), ("lai", 6), ("ndmi", 8)):
            data = Reference(ws, min_col=colidx, min_row=1, max_row=len(mensual))
            ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ws.add_chart(ch, "K2")

    # ---- Hoja 4: Fenología ----
    ws = wb.create_sheet("Fenología")
    encabeza(ws, ["Fecha", "Fase estimada", "Estado", "Esperado", "NDVI", "LAI"])
    for i, rr in enumerate(recorrido, 2):
        for j, v in enumerate([rr["fecha"], rr["fase"], rr["estado"],
                               "sí" if rr["esperado"] else "no", rr["ndvi"], rr["lai"]], 1):
            ws.cell(row=i, column=j, value=v)
    anchos(ws, 14, 6); ws.column_dimensions["B"].width = 26; ws.freeze_panes = "A2"

    # ---- Hoja 5: Radar (si hay) ----
    if radar:
        ws = wb.create_sheet("Radar S1")
        cols = ["Fecha", "VV dB", "VH dB", "RVI", "RVI min", "RVI max", "CR dB",
                "Fiabilidad", "Órbita", "n píxeles"]
        encabeza(ws, cols)
        for i, rr in enumerate(radar, 2):
            for j, v in enumerate([rr["fecha"], rr.get("vv"), rr.get("vh"), rr.get("rvi"),
                                   rr.get("rvi_lo"), rr.get("rvi_hi"), rr.get("cr"),
                                   rr.get("fiabilidad"), rr.get("orbita"), rr.get("n_pixeles")], 1):
                ws.cell(row=i, column=j, value=v)
        anchos(ws, 12, len(cols)); ws.column_dimensions["A"].width = 12; ws.freeze_panes = "A2"
        ch = LineChart(); ch.title = "RVI (Sentinel-1)"; ch.height = 8; ch.width = 18
        data = Reference(ws, min_col=4, min_row=1, max_row=len(radar) + 1)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(radar) + 1))
        ws.add_chart(ch, "L2")

    # ---- Hoja: Clima diario (aqui SI cabe el detalle dia a dia) ----
    if clima and _CLIMA is not None:
        ws = wb.create_sheet("Clima")
        cols = [t for _c, t, _a, _d in _CLIMA.COLUMNAS]
        encabeza(ws, cols)
        for i, dia in enumerate(clima, 2):
            for j, (clave, _t, _a, _d) in enumerate(_CLIMA.COLUMNAS, 1):
                ws.cell(row=i, column=j, value=dia.get(clave))
        anchos(ws, 12, len(cols)); ws.column_dimensions["A"].width = 12
        ws.freeze_panes = "A2"
        ch = LineChart(); ch.title = "Temperatura y lluvia (ERA5-Land, comarca)"
        ch.height = 9; ch.width = 24
        for colidx in (2, 5):                     # T media y lluvia
            ch.add_data(Reference(ws, min_col=colidx, min_row=1, max_row=len(clima) + 1),
                        titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(clima) + 1))
        ws.add_chart(ch, "M2")

    # ---- Hoja: Produccion (bascula, todas las campanas) ----
    if rendimientos:
        ws = wb.create_sheet("Producción")
        cols = ["Campaña", "Tipo", "kg/ha", "Humedad grano %", "Superficie ha",
                "Origen del dato", "Fecha"]
        encabeza(ws, cols)
        for i, r_ in enumerate(rendimientos, 2):
            for j, v in enumerate([r_.get("campana"),
                                   "Siega" if r_.get("tipo") == "SIEGA" else "Cosecha",
                                   r_.get("rendimiento_kg_ha"), r_.get("humedad_grano_pct"),
                                   r_.get("superficie_cosechada_ha"), r_.get("fuente_dato"),
                                   r_.get("fecha")], 1):
                ws.cell(row=i, column=j, value=v)
        anchos(ws, 16, len(cols)); ws.freeze_panes = "A2"

    wb.save(ruta)
